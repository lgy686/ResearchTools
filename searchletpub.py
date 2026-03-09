import os
import re
from datetime import datetime
from urllib.parse import quote, urljoin

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter


BASE = "https://www.letpub.com.cn"
SEARCH_URL = BASE + "/index.php?page=journalapp&view=search&searchname={query}"

# 保存到 .py 脚本所在目录
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(SCRIPT_DIR, "letpub_journal_results.xlsx")

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Referer": BASE,
    "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
}


class LetPubJournalFetcher:
    def __init__(self, timeout=20):
        self.session = requests.Session()
        self.session.headers.update(HEADERS)
        self.timeout = timeout

    def _get(self, url: str) -> str:
        resp = self.session.get(url, timeout=self.timeout)
        resp.raise_for_status()

        if not resp.encoding or resp.encoding.lower() == "iso-8859-1":
            resp.encoding = resp.apparent_encoding

        return resp.text

    @staticmethod
    def _clean_text(text: str) -> str:
        if text is None:
            return ""
        text = text.replace("\xa0", " ")
        text = text.replace("\u3000", " ")
        text = re.sub(r"[ \t\r\f\v]+", " ", text)
        text = re.sub(r"\n+", "\n", text)
        return text.strip()

    @staticmethod
    def _clean_one_line(text: str) -> str:
        text = LetPubJournalFetcher._clean_text(text)
        text = text.replace("\n", " ")
        text = re.sub(r"\s+", " ", text)
        return text.strip()

    @staticmethod
    def _normalize_name(name: str) -> str:
        name = name.strip().lower()
        name = re.sub(r"[\s\-_/]+", " ", name)
        return name

    @staticmethod
    def _extract_first(patterns, text, flags=re.I | re.S):
        for pattern in patterns:
            m = re.search(pattern, text, flags)
            if m:
                if m.lastindex:
                    return m.group(1).strip()
                return m.group(0).strip()
        return None

    @staticmethod
    def _extract_url_after_label(text: str, label: str):
        pattern = rf"{re.escape(label)}\s*[:：]?\s*(https?://[^\s<>\"'）)\]]+)"
        m = re.search(pattern, text, re.I)
        if m:
            return m.group(1).strip()
        return None

    def search_journal_detail_url(self, journal_name: str) -> str:
        url = SEARCH_URL.format(query=quote(journal_name))
        html = self._get(url)
        soup = BeautifulSoup(html, "lxml")

        detail_links = []

        # 方式1：从 a 标签中找详情页
        for a in soup.find_all("a", href=True):
            href = a["href"]
            if "page=journalapp" in href and "view=detail" in href and "journalid=" in href:
                full_url = urljoin(BASE, href)
                title = self._clean_one_line(a.get_text(" ", strip=True))
                detail_links.append((title, full_url))

        # 方式2：直接正则提取详情页链接，增强兼容性
        pattern = r'(?:href=["\']?)(/index\.php\?[^"\']*page=journalapp[^"\']*view=detail[^"\']*journalid=\d+[^"\']*)'
        matches = re.findall(pattern, html, re.I)
        for href in matches:
            full_url = urljoin(BASE, href)
            detail_links.append(("", full_url))

        # 去重
        unique_links = []
        seen = set()
        for title, full_url in detail_links:
            jid_match = re.search(r"journalid=(\d+)", full_url)
            key = jid_match.group(1) if jid_match else full_url
            if key not in seen:
                seen.add(key)
                unique_links.append((title, full_url))

        if not unique_links:
            raise ValueError("未找到期刊详情页链接，可能是搜索词不匹配或网页结构发生变化。")

        target = self._normalize_name(journal_name)

        # 1. 完全匹配
        for title, full_url in unique_links:
            if title and self._normalize_name(title) == target:
                return full_url

        # 2. 包含匹配
        for title, full_url in unique_links:
            if title and target in self._normalize_name(title):
                return full_url

        # 3. 反向包含匹配
        for title, full_url in unique_links:
            if title and self._normalize_name(title) in target:
                return full_url

        # 4. 默认返回第一个
        return unique_links[0][1]

    def parse_detail_page(self, detail_url: str, user_input_name: str) -> dict:
        html = self._get(detail_url)
        soup = BeautifulSoup(html, "lxml")

        raw_text = soup.get_text("\n", strip=True)
        text = self._clean_text(raw_text)
        one_line_text = self._clean_one_line(raw_text)

        title = user_input_name.strip()

        five_year_if = self._extract_first([
            r"五年影响因子\s*([0-9.]+)",
        ], one_line_text)

        oa_status = self._extract_first([
            r"是否OA开放访问\s*(Yes|No)",
            r"是否OA开放访问\s*(是|否)",
        ], one_line_text)

        review_speed_official = self._extract_first([
            r"平均审稿速度\s*期刊官网数据[:：]?\s*(.*?)(?=网友分享经验|平均录用比例|期刊投稿网址|作者指南网址|编辑信息|期刊常用信息链接|年文章数|$)",
            r"平均审稿速度\s*(平均[0-9.]+天)(?=网友分享经验|平均录用比例|期刊投稿网址|作者指南网址|编辑信息|期刊常用信息链接|年文章数|$)",
        ], one_line_text)

        review_speed_user = self._extract_first([
            r"网友分享经验[:：]?\s*(.*?)(?=平均录用比例|期刊投稿网址|作者指南网址|编辑信息|期刊常用信息链接|年文章数|$)",
        ], one_line_text)

        acceptance_rate = self._extract_first([
            r"平均录用比例\s*网友分享经验[:：]?\s*([0-9.]+%)",
            r"平均录用比例\s*([0-9.]+%)",
        ], one_line_text)

        annual_articles = self._extract_first([
            r"年文章数\s*([0-9,]+)",
        ], one_line_text)

        submission_url = self._extract_url_after_label(one_line_text, "期刊投稿网址")
        author_guidelines_url = self._extract_url_after_label(one_line_text, "作者指南网址")

        # 备用提取：从 a 标签中再找
        if not submission_url or not author_guidelines_url:
            all_links = []
            for a in soup.find_all("a", href=True):
                href = a.get("href", "").strip()
                text_a = self._clean_one_line(a.get_text(" ", strip=True))
                if href.startswith("http://") or href.startswith("https://"):
                    all_links.append((text_a, href))

            if not submission_url:
                for txt, href in all_links:
                    low = txt.lower()
                    if "投稿" in txt or "submission" in low or "editorial manager" in low:
                        submission_url = href
                        break

            if not author_guidelines_url:
                for txt, href in all_links:
                    low = txt.lower()
                    if "作者指南" in txt or "guide for authors" in low or "submission guidelines" in low or "instructions for authors" in low:
                        author_guidelines_url = href
                        break

        result = {
            "query_name": user_input_name.strip(),
            "journal_title": title or "未识别",
            "detail_url": detail_url,
            "five_year_impact_factor": five_year_if or "未找到",
            "is_oa": oa_status or "未找到",
            "review_speed_official": review_speed_official or "未找到",
            "review_speed_user_shared": review_speed_user or "未找到",
            "acceptance_rate": acceptance_rate or "未找到",
            "annual_articles": annual_articles or "未找到",
            "submission_url": submission_url or "未找到",
            "author_guidelines_url": author_guidelines_url or "未找到",
            "query_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }

        return result

    def fetch(self, journal_name: str) -> dict:
        detail_url = self.search_journal_detail_url(journal_name)
        return self.parse_detail_page(detail_url, journal_name)


def pretty_print(data: dict):
    print("\n" + "=" * 90)
    print(f"查询名称             : {data.get('query_name')}")
    print(f"期刊名称             : {data.get('journal_title')}")
    print(f"详情页链接           : {data.get('detail_url')}")
    print("-" * 90)
    print(f"五年影响因子         : {data.get('five_year_impact_factor')}")
    print(f"是否 OA              : {data.get('is_oa')}")
    print(f"平均审稿速度(官网)   : {data.get('review_speed_official')}")
    print(f"平均审稿速度(网友)   : {data.get('review_speed_user_shared')}")
    print(f"平均录用比例         : {data.get('acceptance_rate')}")
    print(f"年文章数             : {data.get('annual_articles')}")
    print(f"投稿网址             : {data.get('submission_url')}")
    print(f"作者指南网址         : {data.get('author_guidelines_url')}")
    print(f"查询时间             : {data.get('query_time')}")
    print("=" * 90 + "\n")


def get_display_width(value) -> int:
    """
    估算 Excel 显示宽度：
    - 英文/数字约算 1
    - 中文/全角约算 2
    """
    if value is None:
        return 0

    s = str(value)
    width = 0
    for ch in s:
        if ord(ch) > 127:
            width += 2
        else:
            width += 1
    return width


def auto_adjust_column_width(ws, min_width=8, max_width=80):
    """
    根据内容自动调整列宽
    """
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell_value = cell.value

            if cell_value is None:
                continue

            lines = str(cell_value).split("\n")
            line_max = max(get_display_width(line) for line in lines) if lines else 0

            if line_max > max_len:
                max_len = line_max

        adjusted_width = max(min_width, min(max_len + 2, max_width))
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = adjusted_width


def apply_excel_style(ws):
    """
    Excel 美化：
    1. 表头加粗
    2. 全部内容居中
    3. 自动换行
    4. 加边框
    5. 冻结首行
    6. 自动筛选
    """
    thin_side = Side(style="thin", color="000000")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    header_font = Font(bold=True, size=11)
    body_font = Font(size=10)

    center_alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )

    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and str(cell.value).strip() != "":
                cell.border = border
                cell.alignment = center_alignment

                if cell.row == 1:
                    cell.font = header_font
                else:
                    cell.font = body_font

    # 表头行高
    ws.row_dimensions[1].height = 24

    # 数据行行高
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 20

    # 冻结首行
    ws.freeze_panes = "A2"

    # 自动筛选
    ws.auto_filter.ref = ws.dimensions


def refresh_serial_numbers(ws):
    """
    刷新首列序号：
    第1行为表头，从第2行开始编号 1,2,3,...
    """
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=1, value=row_idx - 1)


def save_to_excel(data: dict, excel_file: str = EXCEL_FILE):
    headers = [
        "序号",
        "查询名称",
        "期刊名称",
        "详情页链接",
        "五年影响因子",
        "是否OA",
        "平均审稿速度(官网)",
        "平均审稿速度(网友)",
        "平均录用比例",
        "年文章数",
        "投稿网址",
        "作者指南网址",
        "查询时间",
    ]

    # 注意：序号先占位，后面统一刷新
    row_data = [
        "",
        data.get("query_name", ""),
        data.get("journal_title", ""),
        data.get("detail_url", ""),
        data.get("five_year_impact_factor", ""),
        data.get("is_oa", ""),
        data.get("review_speed_official", ""),
        data.get("review_speed_user_shared", ""),
        data.get("acceptance_rate", ""),
        data.get("annual_articles", ""),
        data.get("submission_url", ""),
        data.get("author_guidelines_url", ""),
        data.get("query_time", ""),
    ]

    if not os.path.exists(excel_file):
        wb = Workbook()
        ws = wb.active
        ws.title = "LetPub期刊信息"
        ws.append(headers)
        ws.append(row_data)

        refresh_serial_numbers(ws)
        apply_excel_style(ws)
        auto_adjust_column_width(ws)

        wb.save(excel_file)
        print(f"已创建 Excel 文件，并写入首条记录：{excel_file}")
        return

    wb = load_workbook(excel_file)
    ws = wb.active

    current_headers = [ws.cell(row=1, column=i).value for i in range(1, len(headers) + 1)]
    if current_headers != headers:
        print("检测到旧版 Excel 表头结构不一致。")
        print("建议删除旧文件 letpub_journal_results.xlsx 后重新运行，以生成新版表头。")

    detail_url = str(data.get("detail_url", "")).strip()
    found_row = None

    # 现在第4列才是详情页链接（因为第1列加了序号）
    detail_url_col = 4

    for row in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=detail_url_col).value
        if cell_value and str(cell_value).strip() == detail_url:
            found_row = row
            break

    if found_row:
        for col, value in enumerate(row_data, start=1):
            ws.cell(row=found_row, column=col, value=value)
        print(f"已更新已有期刊记录：{data.get('journal_title')}")
    else:
        ws.append(row_data)
        print(f"已追加新期刊记录：{data.get('journal_title')}")

    # 刷新序号
    refresh_serial_numbers(ws)

    # 美化
    apply_excel_style(ws)
    auto_adjust_column_width(ws)

    wb.save(excel_file)
    print(f"Excel 已保存：{excel_file}")


def main():
    fetcher = LetPubJournalFetcher()

    print("LetPub 期刊信息查询工具（增强版：序号 + 冻结首行 + 自动筛选 + Excel美化）")
    print(f"Excel 保存位置：{EXCEL_FILE}")
    print("输入期刊名称进行查询，输入 q 退出。\n")

    while True:
        journal_name = input("请输入期刊名称：").strip()

        if journal_name.lower() == "q":
            print("程序已退出。")
            break

        if not journal_name:
            print("期刊名称不能为空，请重新输入。\n")
            continue

        try:
            result = fetcher.fetch(journal_name)
            pretty_print(result)
            save_to_excel(result, EXCEL_FILE)
            print()
        except requests.HTTPError as e:
            print(f"HTTP 请求失败：{e}\n")
        except Exception as e:
            print(f"查询失败：{e}\n")


if __name__ == "__main__":
    main()